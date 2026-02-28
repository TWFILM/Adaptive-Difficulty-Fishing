import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT_FILE = "calculated_results.xlsx"
OUTPUT_FILE = "Perception_by_Player_Level.xlsx"


def main():

    df = pd.read_excel(INPUT_FILE, engine="openpyxl")

    df_selected = df[
        [
            "Player_Level",
            "Mode_Played",
            "Difficulty_Score",
            "Fun_Score",
            "DDA_Similarity",
            "DDA_More_Fun"
        ]
    ]

    result = (
        df_selected
        .groupby(["Player_Level", "Mode_Played"])
        .mean()
        .reset_index()
    )
    # group by Level + Mode
    result = result.rename(columns={
        "Difficulty_Score": "Avg_Difficulty",
        "Fun_Score": "Avg_Fun",
        "DDA_Similarity": "Avg_DDA_Similarity",
        "DDA_More_Fun": "Avg_DDA_More_Fun"
    })

    result.to_excel(OUTPUT_FILE, index=False)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # หา column ของ Player_Level
    level_col = 1  # เพราะ groupby แล้ว Level จะเป็น column แรก

    for row in range(2, ws.max_row + 1):

        level_value = ws.cell(row=row, column=level_col).value

        if level_value == "Advanced":
            fill_color = green_fill
        elif level_value == "Intermediate":
            fill_color = blue_fill
        elif level_value == "Beginner":
            fill_color = red_fill
        else:
            continue

        # fill color
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill_color

    wb.save(OUTPUT_FILE)

    print("🎨 File saved with colored rows!")


if __name__ == "__main__":
    main()