import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

INPUT_FILE = "results.xlsx"
OUTPUT_FILE = "calculated_results.xlsx"

DIFFICULTY_WEIGHTS = {
    "EASY": 1,
    "MEDIUM": 2,
    "HARD": 3,
    "DDA": 3
}

ACCURACY_WEIGHT = 0.7
WIN_WEIGHT = 0.3

OPTIMAL_TIME = 8.25  # theoretical perfect time (no miss)

def compute_player_skill(df):

    skill_map = {}

    for player in df["Player_ID"].unique():

        player_data = df[df["Player_ID"] == player]

        total_weight = 0
        weighted_sum = 0

        for mode, weight in DIFFICULTY_WEIGHTS.items():

            mode_data = player_data[player_data["Mode_Played"] == mode]
            # ถ้าไม่มีอะไรผิดพลาด len(mode_data) จะเป็น 1 เสมอ คือ 1 คนเล่น โหมดละครั้ง
            if len(mode_data) == 0:
                continue

            row = mode_data.iloc[0]

            accuracy = row["Match_Accuracy"]
            duration = row["Catch_Duration_Sec"]
            win = 1 if row["Win_Loss"] == "WIN" else 0

            # 8.25 / complete minigame time
            speed = OPTIMAL_TIME / duration
            speed = min(speed, 1.0)   # cap at 1
            speed = max(speed, 0.0)   # safety

            mode_score = (
                (accuracy * 0.6) +
                (speed * 0.2) +
                (win * 0.2)
            )

            weighted_sum += mode_score * weight
            total_weight += weight

        if total_weight == 0:
            skill = 0.5
        else:
            skill = weighted_sum / total_weight

        skill = np.clip(skill, 0.1, 0.95)

        skill_map[player] = round(skill, 4)

    return skill_map


def classify_by_quartile(df):

    Q1 = df["Computed_Skill"].quantile(0.25)
    Q3 = df["Computed_Skill"].quantile(0.75)

    def assign_level(skill):
        if skill <= Q1:
            return "Beginner"
        elif skill >= Q3:
            return "Advanced"
        else:
            return "Intermediate"

    df["Player_Level"] = df["Computed_Skill"].apply(assign_level)

    print(f"Q1 (25%): {round(Q1,4)}")
    print(f"Q3 (75%): {round(Q3,4)}")

    return df


if __name__ == "__main__":

    df = pd.read_excel(INPUT_FILE, engine="openpyxl")

    skill_map = compute_player_skill(df)

    df["Computed_Skill"] = df["Player_ID"].map(skill_map)

    df = classify_by_quartile(df)

    df.to_excel(OUTPUT_FILE, index=False)

    print("✅ File saved as:", OUTPUT_FILE)

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # find the column index for "Player_Level"
    header_row = 1
    level_col = None

    for col in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value == "Player_Level":
            level_col = col
            break

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # filled color
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=level_col)
        
        if cell.value == "Advanced":
            cell.fill = green_fill
        elif cell.value == "Intermediate":
            cell.fill = blue_fill
        elif cell.value == "Beginner":
            cell.fill = red_fill

    wb.save(OUTPUT_FILE)

    print("🎨 Added colors to Player_Level column!")